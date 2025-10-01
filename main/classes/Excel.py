
import openpyxl
from openpyxl.styles import Font, Color 
from main.classes.Match import loadMatchesFromCsv
from openpyxl.styles import PatternFill

import os

green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
TEAM_COLUMN = 2

#sets up the text on the left side (so not the actual data for the players)
def setupNonPlayerText(ws):
    row = 1
    column = 1
    #group matches
    matches = loadMatchesFromCsv()
    ws.cell(row=row,column=column,value="Gruppe:")
    ws.cell(row=row,column=column+1,value="Kamp:")
    for match in matches:
        row += 1
        ws.cell(row=row,column=column,value=match.group)
        ws.cell(row=row,column=column+1,value=match.excelPrint())
    #add subtotal row
    row = addSubTotalRowHeader(ws,row,column)
    #ro16 teams:
    ws.cell(row=row,column=column,value="Hold i ottendedelsfinalen:")
    for i in range(16):
        row += 1
        ws.cell(row=row,column=column+1,value=f"Hold {i+1}")
    #add subtotal row
    row = addSubTotalRowHeader(ws,row,column)
    #ro8 teams:
    ws.cell(row=row,column=column,value="Hold i kvartfinalen:")
    for i in range(8):
        row += 1
        ws.cell(row=row,column=column+1,value=f"Hold {i+1}")
    #add subtotal row
    row = addSubTotalRowHeader(ws,row,column)
    #semi final teams:
    ws.cell(row=row,column=column,value="Hold i semifinalen:")
    for i in range(4):
        row += 1
        ws.cell(row=row,column=column+1,value=f"Hold {i+1}")
    #add subtotal row
    row = addSubTotalRowHeader(ws,row,column)
    #finals teams:
    ws.cell(row=row,column=column,value="Hold i finalen:")
    for i in range(2):
        row += 1
        ws.cell(row=row,column=column+1,value=f"Hold {i+1}")
    #winner
    row += 1
    ws.cell(row=row,column=column,value="Vinder:")
    #add subtotal row
    row = addSubTotalRowHeader(ws,row,column)
    #How far Denmark goes
    ws.cell(row=row,column=column,value="Hvor langt når Danmark:")
    row += 1
    #topscorer
    ws.cell(row=row,column=column,value="Topscorer:")
    row+=1
    #dane to score
    ws.cell(row=row,column=column,value="Dansker der scorer:")
    row += 1
    #how many goals by dane
    ws.cell(row=row,column=column,value="Hvor mange mål scorer den valgte dansker:")
    #player to get red card
    row += 1
    ws.cell(row=row,column=column,value="Spiller der får rødt kort:")
    #add subtotal row
    row = addSubTotalRowHeader(ws,row,column)
    #add total row
    row += 1
    ws.cell(row=row,column=column,value="Total:")

def setupPlayerText(ws,player,column):
    row = 1
    ws.cell(row=row,column=column,value=player.name)
    for match in player.groupMatchGames:
        row += 1
        ws.cell(row=row,column=column,value=match)
    row += 2 #subtotal row
    for team in player.ro16Teams:
        row += 1
        ws.cell(row=row,column=column,value=team)
    row += 2 #subtotal row
    for team in player.quarterFinalTeams:
        row += 1
        ws.cell(row=row,column=column,value=team)
    row += 2 #subtotal row
    for team in player.semiFinalTeams:
        row += 1
        ws.cell(row=row,column=column,value=team)
    row += 2 #subtotal row
    for team in player.finalTeams:
        row += 1
        ws.cell(row=row,column=column,value=team)
    row += 1
    ws.cell(row=row,column=column,value=player.winner)
    row += 2 #subtotal row
    ws.cell(row=row,column=column,value=player.howFarDenmarkReaches)
    row += 1
    ws.cell(row=row,column=column,value=player.topScorer.title())
    row += 1
    ws.cell(row=row,column=column,value=player.daneToScore.title())
    row += 1
    ws.cell(row=row,column=column,value=player.howManyGoalsByDane)
    row += 1
    ws.cell(row=row,column=column,value=player.redCardedPlayer.title())
    
        

def setCellWidths(ws):
    column = 1
    
    cellWidths = []
    while True:
        if ws.cell(row=1,column=column).value: #if something is in this column
            longestStringLength = -1
            for row in range(ws.max_row):
                cellValue = ws.cell(row=row+1,column=column).value
                if cellValue:
                    cellValueLength = len(ws.cell(row=row+1,column=column).value)
                    if cellValueLength > longestStringLength:
                        longestStringLength = cellValueLength      
            cellWidths.append(longestStringLength)
            column += 1
        else:
            break

    for i, cellWidth in enumerate(cellWidths):
        ws.column_dimensions[numberToExcelColumn(i+1)].width = cellWidth
    
    
#for excelfile
#number is always at least 1
#max working output = ZZ
def numberToExcelColumn(number):
    result = ""
    chars = " ABCDEFGHIJKLMNOPQRSTUVWXYZ" #first blank space is intended
    while number > len(chars)-1:
        if result != "":
            nextLetterIndex = chars.index(result[-1])+1
            if nextLetterIndex == len(chars):
                nextLetterIndex = 1
            result = result[1:-1] + chars[nextLetterIndex]
            number -= len(chars)-1
        else:
            result += chars[1]
            number -= len(chars)-1
    if (number != 0):
        result += chars[number]
    return result

def setupExcelFile(players):
    file_path = os.path.join('main', 'data', 'slutrundeKasse.xlsx')
    
    if os.path.exists(file_path):
        print(f"Deleting existing file: {file_path}")
        os.remove(file_path)
    else:
        print(f"No existing file to delete: {file_path}")

    
    ##
    wb = openpyxl.Workbook() # new workbook
    ws = wb.active #new worksheet
    ws.title = "Feriekasse"
    setupNonPlayerText(ws)
    column = 3
    for player in players:
        setupPlayerText(ws,player,column)
        column += 1
    #add "result" column
    cell = ws.cell(row=1,column=column,value="Resultat")
    cell.font = Font(bold=True)
    setCellWidths(ws)
    wb.save(fr"main/data/slutrundeKasse.xlsx")
    wb.close()

def updateExcelFile(groupStageMatches, teamsInRo16, teamsInRo8, teamsInSemiFinals, teamsInFinal, winner):
    file_path = os.path.join('main', 'data', 'slutrundeKasse.xlsx')
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # or wb['SheetName'] for a specific sheet
    numOfPlayers = getNumOfPlayers(ws)
    row = 1
    resultColumn = TEAM_COLUMN + numOfPlayers
    if groupStageMatches != []:
        row = handleGroupStageMatches(groupStageMatches, ws, resultColumn, row)
    print(row)
    if teamsInRo16 != []:
        row = handleKnockoutStageMatches(teamsInRo16, ws, resultColumn, row)
    if teamsInRo8 != []:
        row = handleKnockoutStageMatches(teamsInRo8, ws, resultColumn, row)
    if teamsInSemiFinals != []:
        row = handleKnockoutStageMatches(teamsInSemiFinals, ws, resultColumn, row)
    if teamsInFinal != []:
        row = handleFinaleTeams(teamsInFinal, ws, resultColumn, row)
    if winner != None:
        row = handleWinner(winner, ws, resultColumn, row)

    
    setSubtotals(ws, resultColumn)
    setTotal(ws, resultColumn)

    wb.save(fr"main/data/slutrundeKasse.xlsx")
    wb.close()

def getNumOfPlayers(ws):
    column = 2
    while ws.cell(row=1,column=column).value != "Resultat":
        column += 1
    return column - 2 #not counting the first two columns (group and match)

def handleGroupStageMatches(groupStageMatches, ws, resultColumn, row):
    index = 0
    print(len(groupStageMatches))
    while True:
        if index >= len(groupStageMatches):
            break
        print(groupStageMatches[index])
        cell_value = ws.cell(row=row, column=TEAM_COLUMN).value
        if cell_value is None:
            break
        if not (groupStageMatches[index]['home'] in cell_value and groupStageMatches[index]['away'] in cell_value): #not correct match
            row += 1
            continue
        ws.cell(row=row, column=resultColumn, value=groupStageMatches[index]['result'])

        currentColumn = TEAM_COLUMN + 1
        while currentColumn < resultColumn:
            playerCell = ws.cell(row=row, column=currentColumn)
            playerCellValue = playerCell.value
            if playerCellValue and groupStageMatches[index]['result'] in playerCellValue: #player predicted correctly
                playerCell.fill = green_fill
            else:
                playerCell.fill = red_fill
            currentColumn += 1
        index += 1
    return row

#only works til og med semi finals
def handleKnockoutStageMatches(teams, ws, resultColumn, row):
    #subtotal from previous section
    while ws.cell(row=row, column=1).value != "Subtotal:":
        row += 1
    row += 2 #move to start of this knockout section

    originalRow = row

    #farv grøn for alle spillerne i denne sektion
    column = TEAM_COLUMN + 1
    while column < resultColumn:
        row = originalRow
        while ws.cell(row=row, column=column).value is not None:
            cell = ws.cell(row=row, column=column)
            if cell.value in teams:
                cell.fill = green_fill
            row += 1
        column += 1

    #reset row for inserting teams
    row = originalRow

    while True:
        if ws.cell(row=row, column=1).value == "Subtotal:":
            break
        cell = ws.cell(row=row, column=resultColumn)
        #cell not empty
        if cell.value is not None:
            #holdet i cellen er i teams listen: fjern holdet fra listen
            if cell.value in teams:
                teams.remove(cell.value)
            #holdet i cellen er ikke i teams listen: gå videre til næste række
        # ingen hold i cellen: indsæt det første hold i teams listen
        else:
            cell.value = teams.pop(0)
        row += 1
    
    #reset row
    row = originalRow

    #if all teams have been inserted, color the cells red for the wrong teams for each player
    column = TEAM_COLUMN + 1
    while column < resultColumn:
        row = originalRow
        while ws.cell(row=row, column=1).value != "Subtotal:":
            cell = ws.cell(row=row, column=column)
            if cell.fill != green_fill: #only color red if not already colored green
                cell.fill = red_fill
            row += 1
        column += 1
            

    return row

def handleFinaleTeams(teams, ws, resultColumn, row):
    row += 2 #from subtotal to first team in final
    column = TEAM_COLUMN + 1 #the column of the first player

    while column < resultColumn:
        team1 = ws.cell(row=row, column=column).value
        team2 = ws.cell(row=row+1, column=column).value
        if team1 in teams:
            ws.cell(row=row, column=column).fill = green_fill
        else:
            ws.cell(row=row, column=column).fill = red_fill
        if team2 in teams:
            ws.cell(row=row+1, column=column).fill = green_fill
        else:
            ws.cell(row=row+1, column=column).fill = red_fill
        column += 1
    
    cell = ws.cell(row=row, column=resultColumn)
    if cell.value is None:
        cell.value = teams.pop(0)
    else:
        if cell.value in teams:
            teams.remove(cell.value)
    row += 1
    cell = ws.cell(row=row, column=resultColumn)
    if len(teams) > 0:
        if cell.value is None:
            cell.value = teams.pop(0)
        else:
            if cell.value in teams:
                teams.remove(cell.value) 

    row += 1 #now at winner teams row
    return row

def handleWinner(winner, ws, resultColumn, row):
    column = TEAM_COLUMN + 1 #the column of the first player
    while column < resultColumn:
        cell = ws.cell(row=row, column=column)
        if cell.value == winner:
            cell.fill = green_fill
        else:
            cell.fill = red_fill
        column += 1

    ws.cell(row=row, column=resultColumn, value=winner)
    
    row += 1 #move to subtotal row
    return row


def setSubtotals(ws, resultColumn):
    firstColumn = 1
    column = TEAM_COLUMN + 1
    row = 1
    pointsTally=0
    pointsToAward = 1 #change to however many points per correct guess

    while column < resultColumn:
        while ws.cell(row=row, column=firstColumn).value != "Total:":
            cell = ws.cell(row=row, column=column)
            #set points to award based on section of the sheet we are in
            pointsToAward = setPointsToAward(ws.cell(row=row, column=firstColumn).value, pointsToAward)
            #tally points for this player in this section in subtotal
            if ws.cell(row=row, column=firstColumn).value == "Subtotal:":
                ws.cell(row=row, column=column, value=pointsTally)
                pointsTally = 0 #reset for next section
            #If cell is green, they guessed correctly and get points
            if cell.fill == green_fill:
                pointsTally += pointsToAward
            row += 1
        row = 1 #reset row for next player
        pointsToAward = 1 #reset points to award for next player
        pointsTally = 0 #reset points tally for next player
        column += 1

def setPointsToAward(sectionName, pointsToAward):
    if sectionName == "Gruppe:":
        return 1
    elif sectionName == "Hold i ottendedelsfinalen:":
        return 1
    elif sectionName == "Hold i kvartfinalen:":
        return 1
    elif sectionName == "Hold i semifinalen:":
        return 1
    elif sectionName == "Hold i finalen:":
        return 2
    elif sectionName == "Vinder:":
        return 1
    elif sectionName == "Hvor langt når Danmark:":
        return 2
    elif sectionName == "Topscorer:":
        return 2
    elif sectionName == "Dansker der scorer:":
        return 1
    elif sectionName == "Hvor mange mål scorer den valgte dansker:":
        return 1
    elif sectionName == "Spiller der får rødt kort:":
        return 2
    #if no change in section, use same points to add as before
    else:
        return pointsToAward

def setTotal(ws, resultColumn):
    firstColumn = 1
    column = TEAM_COLUMN + 1
    row = 1
    subTotalTally=0

    while column < resultColumn:
        while ws.cell(row=row, column=firstColumn).value != "Total:":
            if ws.cell(row=row, column=firstColumn).value == "Subtotal:":
                cell = ws.cell(row=row, column=column)
                if cell.value is not None:
                    subTotalTally += cell.value
            row += 1
        ws.cell(row=row, column=column, value=subTotalTally)
        subTotalTally = 0
        column += 1
        row = 1 #reset row for next player

def addSubTotalRowHeader(ws,row,column):
    row += 1
    cell = ws.cell(row=row,column=column,value="Subtotal:")
    # Make the entire row bold (for all columns with content in this row)
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).font = Font(bold=True)
    row += 1
    return row

